"""Generic Excel site loading driven by map_clients.json."""

from __future__ import annotations

from typing import Any

from openpyxl import load_workbook

from map_config import resolve_data_file
from utils import clean_text, property_family


def _column_value(row: tuple, header_index: dict[str, int], column_name: str) -> str:
    if not column_name or column_name not in header_index:
        return ""
    index = header_index[column_name]
    if index >= len(row):
        return ""
    return clean_text(row[index])


def _logical_field_value(
    row: tuple,
    header_index: dict[str, int],
    columns: dict[str, str],
    field: str,
) -> str:
    """Read a logical site field (site_code, address, etc.) from a workbook row."""
    column_name = columns.get(field, "")
    if not column_name:
        return ""
    return _column_value(row, header_index, column_name)


def _find_header_row(
    worksheet,
    search_columns: list[str],
    max_rows: int = 25,
) -> tuple[list[str], int]:
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = [clean_text(value) for value in row]
        if all(column in values for column in search_columns):
            return values, row_number
        if row_number >= max_rows:
            break
    raise ValueError(
        "Could not locate the column header row.\n\n"
        f"Expected columns include:\n" + "\n".join(search_columns)
    )


def _resolve_geocode_address(geocode_address: str, address: str) -> str:
    return geocode_address or address


def _dedupe_key(
    *,
    dedupe_by: str,
    site_code: str,
    site_name: str,
    address: str,
    field_values: dict[str, str],
) -> str | None:
    """
    Return a deduplication key, or None when the row must not participate in deduping.

    Blank dedupe values never collapse unrelated sites together.
    """
    if dedupe_by == "site_code":
        value = site_code
    elif dedupe_by == "address":
        value = address
    elif dedupe_by == "site_name":
        value = site_name
    else:
        value = field_values.get(dedupe_by, "")

    if not value:
        return None

    return value.casefold()


def load_client_sites(client: dict[str, Any]) -> list[dict[str, Any]]:
    workbook_path = resolve_data_file(client["file"])
    if not workbook_path.exists():
        raise FileNotFoundError(
            f"[{client['id']}] Workbook not found: {workbook_path.name}\n\n"
            f"Expected location:\n{workbook_path}\n\n"
            f"Place the file in DATA/ or the project root."
        )

    sheet_name = client["sheet"]
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        raise RuntimeError(
            f"[{client['id']}] Failed to open workbook '{workbook_path.name}': {exc}"
        ) from exc

    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise KeyError(
            f"[{client['id']}] Sheet '{sheet_name}' does not exist in {workbook_path.name}."
        )

    worksheet = workbook[sheet_name]
    header_mode = client.get("header_mode", "first_row")

    try:
        if header_mode == "search":
            search_columns = client.get("header_search_columns", [])
            if not search_columns:
                raise ValueError(
                    f"[{client['id']}] header_mode 'search' requires header_search_columns."
                )
            headers, header_row_number = _find_header_row(worksheet, search_columns)
            print(f"[{client['id']}] Header row found: {header_row_number}")
            row_iterator = worksheet.iter_rows(values_only=True)
            for _ in range(header_row_number):
                next(row_iterator, None)
        else:
            row_iterator = worksheet.iter_rows(values_only=True)
            try:
                first_row = next(row_iterator)
            except StopIteration as exc:
                raise ValueError(f"[{client['id']}] Sheet '{sheet_name}' is empty.") from exc
            headers = [clean_text(value) for value in first_row]

        required = set(client.get("required_columns", []))
        missing = required - set(headers)
        if missing:
            raise KeyError(
                f"[{client['id']}] Missing required columns in '{sheet_name}': "
                + ", ".join(sorted(missing))
            )

        header_index = {name: index for index, name in enumerate(headers) if name}
        columns = client.get("columns", {})
        dedupe_by = client.get("dedupe_by", "site_code")
        required_row_fields = client.get("required_row_fields", [])

        sites: list[dict[str, Any]] = []
        seen_dedupe_keys: set[str] = set()
        duplicates: list[str] = []

        for row in row_iterator:
            raw: dict[str, str] = {}
            for header, index in header_index.items():
                raw[header] = clean_text(row[index]) if index < len(row) else ""

            site_name = _logical_field_value(row, header_index, columns, "site_name")
            site_code = _logical_field_value(row, header_index, columns, "site_code")
            address = _logical_field_value(row, header_index, columns, "address")
            geocode_address = _logical_field_value(row, header_index, columns, "geocode_address")
            geocode_address = _resolve_geocode_address(geocode_address, address)

            if not site_code and not site_name and not address:
                continue

            field_values = {
                "site_code": site_code,
                "site_name": site_name,
                "address": address,
                "geocode_address": geocode_address,
                "postal_code": _logical_field_value(row, header_index, columns, "postal_code"),
                "area": _logical_field_value(row, header_index, columns, "area"),
            }

            skip_row = False
            for field in required_row_fields:
                if not field_values.get(field, ""):
                    label = site_name or site_code or address or "(unnamed row)"
                    print(
                        f"WARNING: [{client['id']}] '{label}' "
                        f"missing required field '{field}' and was skipped."
                    )
                    skip_row = True
                    break
            if skip_row:
                continue

            dedupe_key = _dedupe_key(
                dedupe_by=dedupe_by,
                site_code=site_code,
                site_name=site_name,
                address=address,
                field_values=field_values,
            )

            if dedupe_key is not None:
                if dedupe_key in seen_dedupe_keys:
                    duplicates.append(site_code or address or site_name)
                    continue
                seen_dedupe_keys.add(dedupe_key)

            family_regex = client.get("family_regex", r"^[A-Za-z]+")
            family = ""
            if client.get("group_by_family") and site_code:
                family = property_family(site_code, family_regex)

            service_column = columns.get("service_types", "")
            service_raw = _column_value(row, header_index, service_column)
            service_types = [
                s.strip() for s in service_raw.replace(";", ",").split(",") if s.strip()
            ]

            area = field_values["area"] or "Unknown"

            site: dict[str, Any] = {
                "client_id": client["id"],
                "client_name": client["name"],
                "site_name": site_name,
                "site_code": site_code,
                "family": family,
                "address": address,
                "geocode_address": geocode_address,
                "postal_code": field_values["postal_code"],
                "area": area,
                "service_types": service_types,
                "latitude": None,
                "longitude": None,
                "coordinate_source": "",
                "matched_address": "",
                "marker_type": client.get("marker_type", "circle"),
                "marker_color": client.get("marker_color", "#3388ff"),
                "z_index_offset": client.get("z_index_offset", 0),
                "css_prefix": client.get("css_prefix", client["id"].lower()),
                "group_by_family": client.get("group_by_family", False),
                "detail_zoom": client.get("detail_zoom", 13),
                "metadata": dict(raw),
            }
            sites.append(site)

        if duplicates:
            print()
            print(f"WARNING: [{client['id']}] Duplicate entries ignored:")
            print(", ".join(sorted(set(duplicates))))

        return sites
    finally:
        workbook.close()
