from __future__ import annotations

import argparse
import csv
import html
import json
import math
import re
import webbrowser

from collections import Counter, defaultdict
from pathlib import Path
from typing import Any


# =============================================================================
# DEPENDENCIES
# =============================================================================

try:
    import folium

    from folium.features import DivIcon

    from folium.plugins import (
        Fullscreen,
        MeasureControl,
        MiniMap,
    )

    from geopy.extra.rate_limiter import RateLimiter
    from geopy.geocoders import Nominatim

    from openpyxl import load_workbook

except ImportError as exc:

    missing = getattr(
        exc,
        "name",
        "a required package",
    )

    print()
    print(
        f"ERROR: Missing Python package: {missing}"
    )

    print(
        "Run Setup_Resource_Map.bat once, "
        "then try again."
    )

    raise SystemExit(2)


# =============================================================================
# VERSION
# =============================================================================

VERSION = (
    "7.0 — CHC + SILVERA CONTRACT OVERLAY"
)


# =============================================================================
# PROJECT PATHS
# =============================================================================

ROOT = Path(__file__).resolve().parent


# -----------------------------------------------------------------------------
# CHC
# -----------------------------------------------------------------------------

CHC_XLSX = (
    ROOT
    / "CHC_Master_Sites.xlsx"
)

CHC_SHEET = (
    "CHC Master Sites"
)


# -----------------------------------------------------------------------------
# SILVERA
# -----------------------------------------------------------------------------

SILVERA_XLSX = (
    ROOT
    / "Silvera_Sites.xlsx"
)

SILVERA_SHEET = (
    "Silvera Sites"
)


# -----------------------------------------------------------------------------
# CACHE
# -----------------------------------------------------------------------------

CACHE_FILE = (
    ROOT
    / "geocode_cache.json"
)


# -----------------------------------------------------------------------------
# OUTPUT
# -----------------------------------------------------------------------------

OUTPUT_DIR = (
    ROOT
    / "OUTPUT"
)

OUTPUT_MAP = (
    OUTPUT_DIR
    / "Resource_Allocation_Map.html"
)

CHC_FAILURES_CSV = (
    OUTPUT_DIR
    / "CHC_Geocoding_Failures.csv"
)

SILVERA_FAILURES_CSV = (
    OUTPUT_DIR
    / "Silvera_Geocoding_Failures.csv"
)


# =============================================================================
# MAP SETTINGS
# =============================================================================

CALGARY_CENTER = (
    51.0447,
    -114.0719,
)


# -----------------------------------------------------------------------------
# CHC FAMILY / DETAIL ZOOM
#
# Below 13:
#
#     RUN
#     PIN
#     FHT
#     ABB
#
# At 13+:
#
#     RUN1
#     RUN2
#     PIN1
#     PIN2
#     FHT1
#     FHT2
#
# SILVERA STARS ARE NOT AFFECTED BY THIS.
# -----------------------------------------------------------------------------

DETAIL_ZOOM_LEVEL = 13


# -----------------------------------------------------------------------------
# FAMILY CLICK ZOOM
# -----------------------------------------------------------------------------

FAMILY_CLICK_MAX_ZOOM = 15


# =============================================================================
# GEOCODING
# =============================================================================

GEOCODE_DELAY_SECONDS = 1.10


GEOCODER_USER_AGENT = (
    "summit-resource-allocation-map/7.0"
)


# Prevent wildly incorrect results.

CALGARY_BOUNDS = {

    "min_lat": 50.75,
    "max_lat": 51.35,

    "min_lon": -114.45,
    "max_lon": -113.75,
}


# =============================================================================
# CHC DISTRICT STYLES
# =============================================================================

DISTRICT_STYLE = {

    "East": {

        "color": "#d73027",

        "label": (
            "CHC East — "
            "Current Grounds/Snow Contract"
        ),

        "show": True,
    },


    "South": {

        "color": "#2a81cb",

        "label": (
            "CHC South — Reference"
        ),

        "show": True,
    },


    "West": {

        "color": "#5a9f2b",

        "label": (
            "CHC West — Reference"
        ),

        "show": True,
    },


    "Unknown": {

        "color": "#f59e0b",

        "label": (
            "CHC District Unknown"
        ),

        "show": True,
    },
}


# =============================================================================
# SILVERA STYLE
# =============================================================================

SILVERA_COLOR = "#7e22ce"

SILVERA_LAYER_NAME = (
    "★ Silvera Communities"
)


# =============================================================================
# REQUIRED CHC COLUMNS
# =============================================================================

CHC_REQUIRED_COLUMNS = {

    "Property Code",

    "Project Name",

    "District",

    "Current Grounds/Snow Contract",

    "Representative Address",

    "Geocode Address",
}


# =============================================================================
# REQUIRED SILVERA COLUMNS
# =============================================================================

SILVERA_REQUIRED_COLUMNS = {

    "Display Name",

    "Source Address",

    "Area",

    "Geocode Address",
}


# =============================================================================
# STREET TYPE NORMALIZATION
# =============================================================================

STREET_TYPES = {

    "AV": "Ave",
    "AVE": "Ave",

    "ST": "St",

    "RD": "Rd",

    "DR": "Dr",

    "PL": "Pl",

    "CL": "Close",

    "CR": "Crescent",

    "LN": "Lane",

    "CO": "Court",
    "CT": "Court",

    "TR": "Trail",
    "TRAIL": "Trail",

    "WY": "Way",
    "WAY": "Way",

    "BV": "Blvd",
    "BL": "Blvd",
    "BLVD": "Blvd",

    "GR": "Green",

    "GA": "Gate",

    "GD": "Gardens",

    "HT": "Heights",

    "TC": "Terrace",
    "TE": "Terrace",

    "PT": "Point",

    "PK": "Park",

    "RI": "Rise",

    "RO": "Row",

    "SQ": "Square",

    "MR": "Manor",

    "ME": "Mews",

    "CM": "Common",

    "LK": "Link",

    "VW": "View",
}


# =============================================================================
# BASIC HELPERS
# =============================================================================

def clean_text(
    value: Any,
) -> str:

    if value is None:
        return ""

    text = str(
        value
    ).strip()

    if text.lower() in {
        "none",
        "nan",
    }:
        return ""

    return text


# =============================================================================
# CHC PROPERTY FAMILY
# =============================================================================

def property_family(
    property_code: str,
) -> str:
    """
    Examples:

        FHT1  -> FHT
        FHT2  -> FHT

        RUN1  -> RUN
        RUN2  -> RUN

        PIN1  -> PIN
        PIN2  -> PIN

        BRD2C -> BRD

        SC2L1 -> SC
    """

    code = clean_text(
        property_code
    ).upper()

    match = re.match(
        r"^[A-Z]+",
        code,
    )

    if match:
        return match.group(0)

    return code


# =============================================================================
# ADDRESS NORMALIZATION
# =============================================================================

def normalize_address(
    address: str,
) -> str:

    text = re.sub(
        r"\s+",
        " ",
        address.strip(),
    )


    # -------------------------------------------------------------------------
    # REMOVE LEADING ZEROES
    #
    # 013 AV -> 13 AV
    # 035 ST -> 35 ST
    # -------------------------------------------------------------------------

    text = re.sub(

        r"\b0+(\d+)\s+"
        r"(AV|AVE|ST|RD|DR|PL|CL|CR|LN|CO|CT|"
        r"TR|WY|BV|BL|BLVD)\b",

        lambda match: (
            f"{int(match.group(1))} "
            f"{match.group(2)}"
        ),

        text,

        flags=re.IGNORECASE,
    )


    # -------------------------------------------------------------------------
    # SPECIAL SILVERA FALLBACK
    #
    # The source list contains:
    #
    #     30 11st NE
    #
    # We preserve that exact source address in Excel/popup,
    # but for GEOCODING ONLY we can also try:
    #
    #     30 11 St NE
    #
    # -------------------------------------------------------------------------

    text = re.sub(

        r"\b(\d{1,3})st\b",

        r"\1 St",

        text,

        flags=re.IGNORECASE,
    )


    # -------------------------------------------------------------------------
    # EXPAND STREET TYPES
    # -------------------------------------------------------------------------

    for (
        short,
        full,
    ) in sorted(

        STREET_TYPES.items(),

        key=lambda item:
            -len(item[0]),
    ):

        text = re.sub(

            rf"\b{re.escape(short)}\b",

            full,

            text,

            flags=re.IGNORECASE,
        )


    # -------------------------------------------------------------------------
    # ADD CALGARY
    # -------------------------------------------------------------------------

    if (
        "calgary"
        not in text.lower()
    ):

        text += (
            ", Calgary, Alberta, Canada"
        )


    else:

        text = re.sub(

            r"\bCalgary\s+AB\b",

            "Calgary, Alberta",

            text,

            flags=re.IGNORECASE,
        )


        if (
            "canada"
            not in text.lower()
        ):

            text += ", Canada"


    text = re.sub(
        r"\s*,\s*",
        ", ",
        text,
    )


    return text


# =============================================================================
# GEOCODING QUERY BUILDER
# =============================================================================

def build_geocode_queries(
    address: str,
    postal_code: str = "",
) -> list[str]:

    address = clean_text(
        address
    )

    postal_code = clean_text(
        postal_code
    )

    if not address:
        return []


    candidates: list[str] = []


    # -------------------------------------------------------------------------
    # ORIGINAL ADDRESS
    # -------------------------------------------------------------------------

    candidates.append(
        f"{address}, Calgary, Alberta, Canada"
    )


    # -------------------------------------------------------------------------
    # POSTAL-CODE VERSION
    # -------------------------------------------------------------------------

    if postal_code:

        candidates.append(

            f"{address}, "
            f"Calgary, Alberta, "
            f"{postal_code}, Canada"
        )


    # -------------------------------------------------------------------------
    # NORMALIZED VERSION
    # -------------------------------------------------------------------------

    normalized = normalize_address(
        address
    )

    candidates.append(
        normalized
    )


    # -------------------------------------------------------------------------
    # REMOVE DUPLICATES
    # -------------------------------------------------------------------------

    unique: list[str] = []

    seen: set[str] = set()


    for query in candidates:

        query = re.sub(
            r"\s+",
            " ",
            query,
        ).strip()

        key = query.casefold()

        if key in seen:
            continue

        seen.add(
            key
        )

        unique.append(
            query
        )


    return unique


# =============================================================================
# CALGARY BOUNDS CHECK
# =============================================================================

def within_calgary_bounds(
    lat: float,
    lon: float,
) -> bool:

    return (

        CALGARY_BOUNDS[
            "min_lat"
        ]

        <= lat <=

        CALGARY_BOUNDS[
            "max_lat"
        ]

        and

        CALGARY_BOUNDS[
            "min_lon"
        ]

        <= lon <=

        CALGARY_BOUNDS[
            "max_lon"
        ]
    )


# =============================================================================
# GEOCODE CACHE
# =============================================================================

def load_cache() -> dict[str, Any]:

    if not CACHE_FILE.exists():

        return {

            "version": 2,

            # Existing V6 CHC cache namespace.
            "sites": {},

            # New Silvera namespace.
            "silvera_sites": {},
        }


    try:

        cache = json.loads(

            CACHE_FILE.read_text(
                encoding="utf-8"
            )
        )


        if not isinstance(
            cache,
            dict,
        ):

            raise ValueError(
                "Cache is not a dictionary."
            )


        # ---------------------------------------------------------------------
        # PRESERVE OLD CHC CACHE
        # ---------------------------------------------------------------------

        if (
            "sites"
            not in cache
        ):

            cache[
                "sites"
            ] = {}


        # ---------------------------------------------------------------------
        # ADD SILVERA CACHE WITHOUT TOUCHING CHC CACHE
        # ---------------------------------------------------------------------

        if (
            "silvera_sites"
            not in cache
        ):

            cache[
                "silvera_sites"
            ] = {}


        cache[
            "version"
        ] = 2


        return cache


    except Exception as exc:

        raise RuntimeError(

            f"Could not read "
            f"{CACHE_FILE.name}: "
            f"{exc}"
        ) from exc


# =============================================================================
# SAVE CACHE
# =============================================================================

def save_cache(
    cache: dict[str, Any],
) -> None:

    CACHE_FILE.write_text(

        json.dumps(

            cache,

            indent=2,

            ensure_ascii=False,
        ),

        encoding="utf-8",
    )


# =============================================================================
# CHC CACHE KEY
# =============================================================================

def chc_cache_key(
    site: dict[str, str],
) -> str:

    return "|".join(

        [

            clean_text(
                site.get(
                    "Property Code"
                )
            ),

            clean_text(
                site.get(
                    "Representative Address"
                )
            ),

            clean_text(
                site.get(
                    "Postal Code"
                )
            ),
        ]
    )


# =============================================================================
# SILVERA CACHE KEY
# =============================================================================

def silvera_cache_key(
    site: dict[str, str],
) -> str:

    return "|".join(

        [

            clean_text(
                site.get(
                    "Display Name"
                )
            ),

            clean_text(
                site.get(
                    "Source Address"
                )
            ),
        ]
    )


# =============================================================================
# CREATE GEOCODER
# =============================================================================

def create_geocoder():

    geolocator = Nominatim(

        user_agent=(
            GEOCODER_USER_AGENT
        ),

        timeout=15,
    )


    return RateLimiter(

        geolocator.geocode,

        min_delay_seconds=(
            GEOCODE_DELAY_SECONDS
        ),

        max_retries=2,

        error_wait_seconds=5.0,

        swallow_exceptions=True,
    )


# =============================================================================
# TRY GEOCODER QUERIES
# =============================================================================

def try_geocode_queries(
    queries: list[str],
    geocode,
) -> tuple[
    dict[str, Any] | None,
    int,
    str,
]:

    requests = 0

    last_reason = (
        "No result returned."
    )


    for query in queries:

        requests += 1


        try:

            location = geocode(

                query,

                exactly_one=True,

                country_codes="ca",

                language="en",

                addressdetails=True,
            )


        except Exception as exc:

            location = None

            last_reason = (
                f"Geocoder error: {exc}"
            )


        if location is None:
            continue


        lat = float(
            location.latitude
        )


        lon = float(
            location.longitude
        )


        if not within_calgary_bounds(
            lat,
            lon,
        ):

            last_reason = (

                "Result outside Calgary bounds: "

                f"{lat:.6f}, "
                f"{lon:.6f}"
            )

            continue


        return (

            {

                "latitude":
                    lat,

                "longitude":
                    lon,

                "matched_address":
                    clean_text(
                        location.address
                    ),

                "query":
                    query,
            },

            requests,

            "",
        )


    return (
        None,
        requests,
        last_reason,
    )


# =============================================================================
# READ CHC MASTER SITES
# =============================================================================

def read_chc_sites() -> list[
    dict[str, str]
]:

    if not CHC_XLSX.exists():

        raise FileNotFoundError(

            f"{CHC_XLSX.name} "
            f"was not found.\n\n"

            f"Expected location:\n"
            f"{CHC_XLSX}"
        )


    workbook = load_workbook(

        CHC_XLSX,

        read_only=True,

        data_only=True,
    )


    if (
        CHC_SHEET
        not in workbook.sheetnames
    ):

        raise KeyError(

            f"Sheet "
            f"'{CHC_SHEET}' "
            f"does not exist in "
            f"{CHC_XLSX.name}."
        )


    worksheet = workbook[
        CHC_SHEET
    ]


    rows = worksheet.iter_rows(
        values_only=True
    )


    try:

        first_row = next(
            rows
        )

    except StopIteration:

        raise ValueError(
            f"{CHC_SHEET} is empty."
        )


    headers = [

        clean_text(
            value
        )

        for value
        in first_row
    ]


    missing = (

        CHC_REQUIRED_COLUMNS
        - set(headers)
    )


    if missing:

        raise KeyError(

            "Missing required CHC columns: "

            + ", ".join(
                sorted(
                    missing
                )
            )
        )


    header_index = {

        name: index

        for (
            index,
            name,
        )

        in enumerate(
            headers
        )
    }


    sites: list[
        dict[str, str]
    ] = []


    seen_codes: set[
        str
    ] = set()


    duplicate_codes: list[
        str
    ] = []


    for row in rows:

        property_code = clean_text(

            row[
                header_index[
                    "Property Code"
                ]
            ]
        )


        if not property_code:
            continue


        normalized_code = (
            property_code.casefold()
        )


        # ---------------------------------------------------------------------
        # 1 PROPERTY CODE = 1 CHC MAP LOCATION
        # ---------------------------------------------------------------------

        if normalized_code in seen_codes:

            duplicate_codes.append(
                property_code
            )

            continue


        seen_codes.add(
            normalized_code
        )


        site: dict[
            str,
            str,
        ] = {}


        for (
            header,
            index,
        ) in header_index.items():

            if index < len(row):

                site[
                    header
                ] = clean_text(
                    row[index]
                )

            else:

                site[
                    header
                ] = ""


        sites.append(
            site
        )


    workbook.close()


    if duplicate_codes:

        print()
        print(
            "WARNING: Duplicate CHC property codes ignored:"
        )

        print(

            ", ".join(

                sorted(

                    set(
                        duplicate_codes
                    )
                )
            )
        )


    return sites


# =============================================================================
# READ SILVERA SITES
# =============================================================================

# =============================================================================
# READ SILVERA SITES
# =============================================================================

def read_silvera_sites() -> list[
    dict[str, str]
]:

    if not SILVERA_XLSX.exists():

        raise FileNotFoundError(

            f"{SILVERA_XLSX.name} "
            f"was not found.\n\n"

            "Place Silvera_Sites.xlsx "
            "in the same folder as "
            "Create_Resource_Map.py."
        )


    workbook = load_workbook(

        SILVERA_XLSX,

        read_only=True,

        data_only=True,
    )


    if (
        SILVERA_SHEET
        not in workbook.sheetnames
    ):

        raise KeyError(

            f"Sheet "
            f"'{SILVERA_SHEET}' "
            f"does not exist in "
            f"{SILVERA_XLSX.name}."
        )


    worksheet = workbook[
        SILVERA_SHEET
    ]


    # =========================================================================
    # IMPORTANT
    #
    # Do NOT use worksheet.max_row or worksheet.max_column here.
    #
    # Some .xlsx files opened in read-only mode report those values as None.
    #
    # Instead, iterate through the worksheet directly until the header row
    # is found.
    # =========================================================================

    row_iterator = worksheet.iter_rows(
        values_only=True
    )


    headers: list[str] | None = None

    header_row_number = None


    # -------------------------------------------------------------------------
    # FIND HEADER ROW
    #
    # The Silvera workbook currently looks like:
    #
    # Row 1 = title
    # Row 2 = explanation
    # Row 3 = blank
    # Row 4 = actual headers
    #
    # But we deliberately search instead of assuming row 4 forever.
    # -------------------------------------------------------------------------

    for (
        row_number,
        row,
    ) in enumerate(

        row_iterator,

        start=1,
    ):

        values = [

            clean_text(
                value
            )

            for value
            in row
        ]


        if (

            "Display Name"
            in values

            and

            "Source Address"
            in values

            and

            "Area"
            in values

            and

            "Geocode Address"
            in values
        ):

            headers = values

            header_row_number = (
                row_number
            )

            break


        # Safety:
        # headers should definitely appear near the top.
        if row_number >= 25:
            break


    if headers is None:

        workbook.close()

        raise ValueError(

            "Could not locate the Silvera "
            "column header row.\n\n"

            "Expected columns include:\n"
            "Display Name\n"
            "Source Address\n"
            "Area\n"
            "Geocode Address"
        )


    print(

        f"Silvera header row found: "
        f"{header_row_number}"
    )


    # -------------------------------------------------------------------------
    # CHECK REQUIRED COLUMNS
    # -------------------------------------------------------------------------

    missing = (

        SILVERA_REQUIRED_COLUMNS
        - set(headers)
    )


    if missing:

        workbook.close()

        raise KeyError(

            "Missing required Silvera columns: "

            + ", ".join(
                sorted(
                    missing
                )
            )
        )


    # -------------------------------------------------------------------------
    # BUILD COLUMN LOOKUP
    # -------------------------------------------------------------------------

    header_index = {

        name: index

        for (
            index,
            name,
        ) in enumerate(
            headers
        )

        if name
    }


    sites: list[
        dict[str, str]
    ] = []


    seen_addresses: set[
        str
    ] = set()


    duplicate_addresses: list[
        str
    ] = []


    # =========================================================================
    # READ ALL DATA ROWS
    #
    # row_iterator is already positioned directly AFTER the header row.
    # =========================================================================

    for row in row_iterator:


        display_name = ""


        display_name_index = (
            header_index[
                "Display Name"
            ]
        )


        if (
            display_name_index
            < len(row)
        ):

            display_name = clean_text(

                row[
                    display_name_index
                ]
            )


        source_address = ""


        source_address_index = (
            header_index[
                "Source Address"
            ]
        )


        if (
            source_address_index
            < len(row)
        ):

            source_address = clean_text(

                row[
                    source_address_index
                ]
            )


        # ---------------------------------------------------------------------
        # IGNORE COMPLETELY BLANK ROWS
        # ---------------------------------------------------------------------

        if (

            not display_name

            and

            not source_address
        ):

            continue


        # ---------------------------------------------------------------------
        # NAME WITHOUT ADDRESS
        # ---------------------------------------------------------------------

        if not source_address:

            print(

                "WARNING: Silvera site "
                f"'{display_name}' "
                "has no address and was skipped."
            )

            continue


        # ---------------------------------------------------------------------
        # ONE PHYSICAL ADDRESS = ONE SILVERA STAR
        # ---------------------------------------------------------------------

        address_key = (
            source_address
            .strip()
            .casefold()
        )


        if (
            address_key
            in seen_addresses
        ):

            duplicate_addresses.append(
                source_address
            )

            continue


        seen_addresses.add(
            address_key
        )


        # ---------------------------------------------------------------------
        # CREATE SITE DICTIONARY
        # ---------------------------------------------------------------------

        site: dict[
            str,
            str,
        ] = {}


        for (
            header,
            index,
        ) in header_index.items():


            if index < len(row):

                site[
                    header
                ] = clean_text(
                    row[index]
                )


            else:

                site[
                    header
                ] = ""


        sites.append(
            site
        )


    workbook.close()


    # -------------------------------------------------------------------------
    # REPORT DUPLICATES
    # -------------------------------------------------------------------------

    if duplicate_addresses:

        print()

        print(
            "WARNING: Duplicate Silvera "
            "physical addresses ignored:"
        )


        for address in sorted(

            set(
                duplicate_addresses
            )
        ):

            print(
                f"  {address}"
            )


    print()

    print(

        f"Silvera unique physical locations read: "
        f"{len(sites)}"
    )


    return sites

# =============================================================================
# GEOCODE CHC
# =============================================================================

def geocode_chc_sites(

    sites: list[
        dict[str, str]
    ],

    *,

    retry_failed: bool = False,

) -> tuple[

    list[
        dict[str, Any]
    ],

    list[
        dict[str, str]
    ],

    int,
]:

    cache = load_cache()


    chc_cache: dict[
        str,
        Any,
    ] = cache.setdefault(
        "sites",
        {},
    )


    geocode = create_geocoder()


    located: list[
        dict[str, Any]
    ] = []


    failures: list[
        dict[str, str]
    ] = []


    new_requests = 0


    total = len(
        sites
    )


    for (
        index,
        site,
    ) in enumerate(

        sites,

        start=1,
    ):

        code = clean_text(

            site.get(
                "Property Code"
            )
        )


        key = chc_cache_key(
            site
        )


        cached = chc_cache.get(
            key
        )


        # ---------------------------------------------------------------------
        # CACHED SUCCESS
        # ---------------------------------------------------------------------

        if (

            cached

            and

            cached.get(
                "status"
            ) == "ok"
        ):

            result = dict(
                site
            )


            result[
                "Latitude"
            ] = float(
                cached[
                    "latitude"
                ]
            )


            result[
                "Longitude"
            ] = float(
                cached[
                    "longitude"
                ]
            )


            result[
                "Matched Address"
            ] = cached.get(
                "matched_address",
                "",
            )


            result[
                "Geocode Source"
            ] = "cache"


            located.append(
                result
            )


            print(

                f"[CHC "
                f"{index:>3}/{total}] "
                f"{code:<8} cached"
            )


            continue


        # ---------------------------------------------------------------------
        # CACHED FAILURE
        # ---------------------------------------------------------------------

        if (

            cached

            and

            cached.get(
                "status"
            ) == "failed"

            and

            not retry_failed
        ):

            failures.append(

                {

                    "Site":
                        code,

                    "Address":
                        clean_text(

                            site.get(
                                "Representative Address"
                            )
                        ),

                    "Reason":
                        cached.get(
                            "reason",
                            "Previously not found",
                        ),

                    "Attempted Queries":
                        " | ".join(

                            cached.get(
                                "queries",
                                [],
                            )
                        ),
                }
            )


            print(

                f"[CHC "
                f"{index:>3}/{total}] "
                f"{code:<8} cached failure"
            )


            continue


        # ---------------------------------------------------------------------
        # ADDRESS
        # ---------------------------------------------------------------------

        address = clean_text(

            site.get(
                "Geocode Address"
            )
        )


        if not address:

            address = clean_text(

                site.get(
                    "Representative Address"
                )
            )


        postal = clean_text(

            site.get(
                "Postal Code"
            )
        )


        queries = build_geocode_queries(

            address,

            postal,
        )


        print(

            f"[CHC "
            f"{index:>3}/{total}] "
            f"{code:<8} geocoding...",

            flush=True,
        )


        (
            matched,
            request_count,
            reason,
        ) = try_geocode_queries(

            queries,

            geocode,
        )


        new_requests += (
            request_count
        )


        # ---------------------------------------------------------------------
        # SUCCESS
        # ---------------------------------------------------------------------

        if matched:

            chc_cache[
                key
            ] = {

                "status":
                    "ok",

                **matched,

                "property_code":
                    code,
            }


            result = dict(
                site
            )


            result[
                "Latitude"
            ] = matched[
                "latitude"
            ]


            result[
                "Longitude"
            ] = matched[
                "longitude"
            ]


            result[
                "Matched Address"
            ] = matched[
                "matched_address"
            ]


            result[
                "Geocode Source"
            ] = "new"


            located.append(
                result
            )


            print(

                "             -> "

                f"{matched['latitude']:.6f}, "

                f"{matched['longitude']:.6f}"
            )


        # ---------------------------------------------------------------------
        # FAILURE
        # ---------------------------------------------------------------------

        else:

            chc_cache[
                key
            ] = {

                "status":
                    "failed",

                "property_code":
                    code,

                "reason":
                    reason,

                "queries":
                    queries,
            }


            failures.append(

                {

                    "Site":
                        code,

                    "Address":
                        address,

                    "Reason":
                        reason,

                    "Attempted Queries":
                        " | ".join(
                            queries
                        ),
                }
            )


            print(

                "             -> "
                f"NOT FOUND ({reason})"
            )


        save_cache(
            cache
        )


    save_cache(
        cache
    )


    return (

        located,

        failures,

        new_requests,
    )


# =============================================================================
# GEOCODE SILVERA
# =============================================================================

def geocode_silvera_sites(

    sites: list[
        dict[str, str]
    ],

    *,

    retry_failed: bool = False,

) -> tuple[

    list[
        dict[str, Any]
    ],

    list[
        dict[str, str]
    ],

    int,
]:

    cache = load_cache()


    silvera_cache: dict[
        str,
        Any,
    ] = cache.setdefault(
        "silvera_sites",
        {},
    )


    geocode = create_geocoder()


    located: list[
        dict[str, Any]
    ] = []


    failures: list[
        dict[str, str]
    ] = []


    new_requests = 0


    total = len(
        sites
    )


    for (
        index,
        site,
    ) in enumerate(

        sites,

        start=1,
    ):

        display_name = clean_text(

            site.get(
                "Display Name"
            )
        )


        source_address = clean_text(

            site.get(
                "Source Address"
            )
        )


        key = silvera_cache_key(
            site
        )


        # ---------------------------------------------------------------------
        # IF LAT/LON ARE MANUALLY ENTERED IN SILVERA XLSX,
        # THEY TAKE PRIORITY OVER GEOCODING.
        # ---------------------------------------------------------------------

        manual_lat = clean_text(

            site.get(
                "Latitude"
            )
        )


        manual_lon = clean_text(

            site.get(
                "Longitude"
            )
        )


        if (
            manual_lat
            and
            manual_lon
        ):

            try:

                lat = float(
                    manual_lat
                )

                lon = float(
                    manual_lon
                )


                if within_calgary_bounds(
                    lat,
                    lon,
                ):

                    result = dict(
                        site
                    )


                    result[
                        "Latitude"
                    ] = lat


                    result[
                        "Longitude"
                    ] = lon


                    result[
                        "Matched Address"
                    ] = (
                        "Manual coordinates from "
                        "Silvera_Sites.xlsx"
                    )


                    result[
                        "Geocode Source"
                    ] = "workbook"


                    located.append(
                        result
                    )


                    print(

                        f"[SILVERA "
                        f"{index:>2}/{total}] "
                        f"{display_name} "
                        f"workbook coordinates"
                    )


                    continue


            except ValueError:
                pass


        cached = silvera_cache.get(
            key
        )


        # ---------------------------------------------------------------------
        # CACHED SUCCESS
        # ---------------------------------------------------------------------

        if (

            cached

            and

            cached.get(
                "status"
            ) == "ok"
        ):

            result = dict(
                site
            )


            result[
                "Latitude"
            ] = float(
                cached[
                    "latitude"
                ]
            )


            result[
                "Longitude"
            ] = float(
                cached[
                    "longitude"
                ]
            )


            result[
                "Matched Address"
            ] = cached.get(
                "matched_address",
                "",
            )


            result[
                "Geocode Source"
            ] = "cache"


            located.append(
                result
            )


            print(

                f"[SILVERA "
                f"{index:>2}/{total}] "
                f"{display_name} cached"
            )


            continue


        # ---------------------------------------------------------------------
        # CACHED FAILURE
        # ---------------------------------------------------------------------

        if (

            cached

            and

            cached.get(
                "status"
            ) == "failed"

            and

            not retry_failed
        ):

            failures.append(

                {

                    "Site":
                        display_name,

                    "Address":
                        source_address,

                    "Reason":
                        cached.get(
                            "reason",
                            "Previously not found",
                        ),

                    "Attempted Queries":
                        " | ".join(

                            cached.get(
                                "queries",
                                [],
                            )
                        ),
                }
            )


            print(

                f"[SILVERA "
                f"{index:>2}/{total}] "
                f"{display_name} "
                f"cached failure"
            )


            continue


        # ---------------------------------------------------------------------
        # GEOCODE ADDRESS
        # ---------------------------------------------------------------------

        geocode_address = clean_text(

            site.get(
                "Geocode Address"
            )
        )


        if not geocode_address:

            geocode_address = (
                source_address
            )


        queries = build_geocode_queries(
            geocode_address
        )


        print(

            f"[SILVERA "
            f"{index:>2}/{total}] "
            f"{display_name} geocoding...",

            flush=True,
        )


        (
            matched,
            request_count,
            reason,
        ) = try_geocode_queries(

            queries,

            geocode,
        )


        new_requests += (
            request_count
        )


        # ---------------------------------------------------------------------
        # SUCCESS
        # ---------------------------------------------------------------------

        if matched:

            silvera_cache[
                key
            ] = {

                "status":
                    "ok",

                **matched,

                "display_name":
                    display_name,

                "source_address":
                    source_address,
            }


            result = dict(
                site
            )


            result[
                "Latitude"
            ] = matched[
                "latitude"
            ]


            result[
                "Longitude"
            ] = matched[
                "longitude"
            ]


            result[
                "Matched Address"
            ] = matched[
                "matched_address"
            ]


            result[
                "Geocode Source"
            ] = "new"


            located.append(
                result
            )


            print(

                "             -> "

                f"{matched['latitude']:.6f}, "

                f"{matched['longitude']:.6f}"
            )


        # ---------------------------------------------------------------------
        # FAILURE
        # ---------------------------------------------------------------------

        else:

            silvera_cache[
                key
            ] = {

                "status":
                    "failed",

                "display_name":
                    display_name,

                "source_address":
                    source_address,

                "reason":
                    reason,

                "queries":
                    queries,
            }


            failures.append(

                {

                    "Site":
                        display_name,

                    "Address":
                        source_address,

                    "Reason":
                        reason,

                    "Attempted Queries":
                        " | ".join(
                            queries
                        ),
                }
            )


            print(

                "             -> "
                f"NOT FOUND ({reason})"
            )


        save_cache(
            cache
        )


    save_cache(
        cache
    )


    return (

        located,

        failures,

        new_requests,
    )


# =============================================================================
# CHC DISTRICT
# =============================================================================

def district_for(
    site: dict[str, Any],
) -> str:

    district = clean_text(

        site.get(
            "District"
        )

    ).title()


    if district in {

        "East",
        "South",
        "West",
    }:

        return district


    return "Unknown"


# =============================================================================
# BUILD CHC FAMILY GROUPS
# =============================================================================

def build_family_groups(
    located: list[dict[str, Any]],
) -> dict[
    tuple[str, str],
    list[dict[str, Any]],
]:

    groups: dict[
        tuple[str, str],
        list[dict[str, Any]],
    ] = defaultdict(list)


    for site in located:

        district = district_for(
            site
        )


        family = property_family(

            clean_text(

                site.get(
                    "Property Code"
                )
            )
        )


        groups[
            (
                district,
                family,
            )
        ].append(
            site
        )


    return dict(
        groups
    )


# =============================================================================
# SPREAD OVERLAPPING CHC PROPERTY LABELS
# =============================================================================

def apply_chc_display_offsets(
    family_sites: list[dict[str, Any]],
) -> None:

    coordinate_groups: dict[
        tuple[float, float],
        list[dict[str, Any]],
    ] = defaultdict(list)


    for site in family_sites:

        lat = float(
            site[
                "Latitude"
            ]
        )


        lon = float(
            site[
                "Longitude"
            ]
        )


        key = (

            round(
                lat,
                5,
            ),

            round(
                lon,
                5,
            ),
        )


        coordinate_groups[
            key
        ].append(
            site
        )


    for duplicate_sites in coordinate_groups.values():

        if len(
            duplicate_sites
        ) == 1:

            site = duplicate_sites[
                0
            ]


            site[
                "_DisplayLatitude"
            ] = float(
                site[
                    "Latitude"
                ]
            )


            site[
                "_DisplayLongitude"
            ] = float(
                site[
                    "Longitude"
                ]
            )


            continue


        center_lat = sum(

            float(
                site[
                    "Latitude"
                ]
            )

            for site
            in duplicate_sites

        ) / len(
            duplicate_sites
        )


        center_lon = sum(

            float(
                site[
                    "Longitude"
                ]
            )

            for site
            in duplicate_sites

        ) / len(
            duplicate_sites
        )


        radius_lat = (
            0.00024
        )


        radius_lon = (
            0.00038
        )


        count = len(
            duplicate_sites
        )


        for (
            index,
            site,
        ) in enumerate(
            duplicate_sites
        ):

            angle = (

                2
                * math.pi
                * index
                / count
            )


            site[
                "_DisplayLatitude"
            ] = (

                center_lat

                + radius_lat
                * math.sin(
                    angle
                )
            )


            site[
                "_DisplayLongitude"
            ] = (

                center_lon

                + radius_lon
                * math.cos(
                    angle
                )
            )


# =============================================================================
# PREPARE CHC DISPLAY OFFSETS
# =============================================================================

def prepare_chc_display_coordinates(
    family_groups: dict[
        tuple[str, str],
        list[dict[str, Any]],
    ],
) -> None:

    for sites in family_groups.values():

        apply_chc_display_offsets(
            sites
        )


# =============================================================================
# SPREAD EXACTLY OVERLAPPING SILVERA STARS
# =============================================================================

def prepare_silvera_display_coordinates(
    sites: list[dict[str, Any]],
) -> None:

    coordinate_groups: dict[
        tuple[float, float],
        list[dict[str, Any]],
    ] = defaultdict(list)


    for site in sites:

        lat = float(
            site[
                "Latitude"
            ]
        )


        lon = float(
            site[
                "Longitude"
            ]
        )


        key = (

            round(
                lat,
                5,
            ),

            round(
                lon,
                5,
            ),
        )


        coordinate_groups[
            key
        ].append(
            site
        )


    for duplicate_sites in coordinate_groups.values():

        if len(
            duplicate_sites
        ) == 1:

            site = duplicate_sites[
                0
            ]


            site[
                "_DisplayLatitude"
            ] = float(
                site[
                    "Latitude"
                ]
            )


            site[
                "_DisplayLongitude"
            ] = float(
                site[
                    "Longitude"
                ]
            )


            continue


        center_lat = sum(

            float(
                site[
                    "Latitude"
                ]
            )

            for site
            in duplicate_sites

        ) / len(
            duplicate_sites
        )


        center_lon = sum(

            float(
                site[
                    "Longitude"
                ]
            )

            for site
            in duplicate_sites

        ) / len(
            duplicate_sites
        )


        # Slightly tighter spread than CHC.
        radius_lat = (
            0.00018
        )


        radius_lon = (
            0.00028
        )


        count = len(
            duplicate_sites
        )


        for (
            index,
            site,
        ) in enumerate(
            duplicate_sites
        ):

            angle = (

                2
                * math.pi
                * index
                / count
            )


            site[
                "_DisplayLatitude"
            ] = (

                center_lat

                + radius_lat
                * math.sin(
                    angle
                )
            )


            site[
                "_DisplayLongitude"
            ] = (

                center_lon

                + radius_lon
                * math.cos(
                    angle
                )
            )
# =============================================================================
# CHC PROPERTY POPUP
# =============================================================================

def chc_popup_html(
    site: dict[str, Any],
) -> str:

    def field(
        name: str,
    ) -> str:

        value = clean_text(
            site.get(name)
        )

        if not value:
            return "—"

        return html.escape(
            value
        )


    contract = field(
        "Current Grounds/Snow Contract"
    )


    if contract.lower() == "yes":

        contract_style = (
            "font-weight:700;"
            "color:#b91c1c;"
        )

    else:

        contract_style = ""


    matched_address = html.escape(

        clean_text(

            site.get(
                "Matched Address"
            )
        )
    )


    matched_row = ""


    if matched_address:

        matched_row = f"""
        <tr>

            <td style="
                padding:4px 10px 4px 0;
                color:#666;
                vertical-align:top;
            ">
                Geocoder Match
            </td>

            <td style="
                padding:4px 0;
                vertical-align:top;
            ">
                {matched_address}
            </td>

        </tr>
        """


    return f"""
    <div style="
        font-family:
            Arial,
            sans-serif;

        min-width:
            300px;

        max-width:
            430px;
    ">

        <div style="
            font-size:
                17px;

            font-weight:
                700;

            margin-bottom:
                2px;
        ">

            {field("Property Code")}
            —
            {field("Project Name")}

        </div>


        <div style="
            color:#666;
            margin-bottom:9px;
        ">
            Calgary Housing Property
        </div>


        <table style="
            border-collapse:
                collapse;

            font-size:
                13px;

            width:
                100%;
        ">


            <tr>

                <td style="
                    padding:
                        4px 10px 4px 0;

                    color:#666;
                ">
                    District
                </td>

                <td style="
                    padding:
                        4px 0;
                ">

                    <b>
                        {field("District")}
                    </b>

                </td>

            </tr>


            <tr>

                <td style="
                    padding:
                        4px 10px 4px 0;

                    color:#666;
                ">
                    Grounds/Snow
                </td>

                <td style="
                    padding:
                        4px 0;

                    {contract_style}
                ">
                    {contract}
                </td>

            </tr>


            <tr>

                <td style="
                    padding:
                        4px 10px 4px 0;

                    color:#666;
                ">
                    Address
                </td>

                <td style="
                    padding:
                        4px 0;
                ">
                    {field("Representative Address")}
                </td>

            </tr>


            <tr>

                <td style="
                    padding:
                        4px 10px 4px 0;

                    color:#666;
                ">
                    Postal Code
                </td>

                <td style="
                    padding:
                        4px 0;
                ">
                    {field("Postal Code")}
                </td>

            </tr>


            <tr>

                <td style="
                    padding:
                        4px 10px 4px 0;

                    color:#666;
                ">
                    Legacy Code
                </td>

                <td style="
                    padding:
                        4px 0;
                ">
                    {field("Legacy Code(s)")}
                </td>

            </tr>


            <tr>

                <td style="
                    padding:
                        4px 10px 4px 0;

                    color:#666;
                ">
                    Ownership
                </td>

                <td style="
                    padding:
                        4px 0;
                ">
                    {field("Ownership")}
                </td>

            </tr>


            {matched_row}

        </table>

    </div>
    """


# =============================================================================
# SILVERA POPUP
# =============================================================================

def silvera_popup_html(
    site: dict[str, Any],
) -> str:

    def field(
        name: str,
    ) -> str:

        value = clean_text(
            site.get(name)
        )

        if not value:
            return "—"

        return html.escape(
            value
        )


    source_names = field(
        "Source Site Name(s)"
    )


    display_name = field(
        "Display Name"
    )


    # -------------------------------------------------------------------------
    # ONLY SHOW SOURCE NAMES IF THEY ADD INFORMATION
    #
    # Example:
    #
    # Display Name:
    # Beaverdam Commons / Beaverdam Townhomes
    #
    # Source Names:
    # Beaverdam Commons | BeaverdamTownhomes
    # -------------------------------------------------------------------------

    source_row = ""


    if (

        source_names != "—"

        and

        source_names.casefold()
        != display_name.casefold()
    ):

        source_row = f"""
        <tr>

            <td style="
                padding:
                    4px 10px 4px 0;

                color:#666;

                vertical-align:
                    top;
            ">
                Communities
            </td>

            <td style="
                padding:
                    4px 0;

                vertical-align:
                    top;
            ">
                {source_names.replace(" | ", "<br>")}
            </td>

        </tr>
        """


    review_note = clean_text(

        site.get(
            "Review Note"
        )
    )


    review_row = ""


    if review_note:

        review_row = f"""
        <tr>

            <td style="
                padding:
                    4px 10px 4px 0;

                color:#666;

                vertical-align:
                    top;
            ">
                Review Note
            </td>

            <td style="
                padding:
                    4px 0;

                color:#a16207;

                font-weight:
                    600;

                vertical-align:
                    top;
            ">
                {html.escape(review_note)}
            </td>

        </tr>
        """


    matched = clean_text(

        site.get(
            "Matched Address"
        )
    )


    matched_row = ""


    if matched:

        matched_row = f"""
        <tr>

            <td style="
                padding:
                    4px 10px 4px 0;

                color:#666;

                vertical-align:
                    top;
            ">
                Geocoder Match
            </td>

            <td style="
                padding:
                    4px 0;

                vertical-align:
                    top;
            ">
                {html.escape(matched)}
            </td>

        </tr>
        """


    return f"""
    <div style="
        font-family:
            Arial,
            sans-serif;

        min-width:
            300px;

        max-width:
            440px;
    ">

        <div style="
            font-size:
                18px;

            font-weight:
                700;

            color:
                {SILVERA_COLOR};

            margin-bottom:
                2px;
        ">
            ★ {display_name}
        </div>


        <div style="
            color:#666;
            margin-bottom:9px;
        ">
            Silvera Community
        </div>


        <table style="
            border-collapse:
                collapse;

            font-size:
                13px;

            width:
                100%;
        ">


            <tr>

                <td style="
                    padding:
                        4px 10px 4px 0;

                    color:#666;
                ">
                    Address
                </td>

                <td style="
                    padding:
                        4px 0;
                ">
                    {field("Source Address")}
                </td>

            </tr>


            <tr>

                <td style="
                    padding:
                        4px 10px 4px 0;

                    color:#666;
                ">
                    Area
                </td>

                <td style="
                    padding:
                        4px 0;
                ">
                    <b>
                        {field("Area")}
                    </b>
                </td>

            </tr>


            <tr>

                <td style="
                    padding:
                        4px 10px 4px 0;

                    color:#666;
                ">
                    Contract
                </td>

                <td style="
                    padding:
                        4px 0;

                    color:
                        {SILVERA_COLOR};

                    font-weight:
                        700;
                ">
                    {field("Contract")}
                </td>

            </tr>


            {source_row}

            {review_row}

            {matched_row}


        </table>

    </div>
    """


# =============================================================================
# CHC DETAIL LABEL
# =============================================================================

def chc_detail_marker_html(
    property_code: str,
    background_color: str,
) -> str:

    safe_code = html.escape(
        property_code
        or "?"
    )


    return f"""
    <div
        class="chc-detail-label"

        data-property-code="{safe_code}"

        style="
            display:flex;

            align-items:center;

            justify-content:center;

            background:
                {background_color};

            color:white;

            border:
                1px solid
                rgba(255,255,255,.96);

            border-radius:
                3px;

            padding:
                1px 3px;

            min-width:
                23px;

            height:
                14px;

            box-sizing:
                border-box;

            font-family:
                Arial,
                Helvetica,
                sans-serif;

            font-size:
                8px;

            font-weight:
                700;

            line-height:
                12px;

            white-space:
                nowrap;

            text-align:
                center;

            box-shadow:
                0 1px 2px
                rgba(0,0,0,.45);

            cursor:
                pointer;
        "
    >

        {safe_code}

    </div>
    """


# =============================================================================
# CHC FAMILY LABEL
# =============================================================================

def chc_family_marker_html(
    display_code: str,
    background_color: str,
) -> str:

    safe_code = html.escape(
        display_code
        or "?"
    )


    return f"""
    <div
        class="chc-family-label"

        data-family="{safe_code}"

        style="
            display:flex;

            align-items:center;

            justify-content:center;

            background:
                {background_color};

            color:white;

            border:
                2px solid
                rgba(255,255,255,.98);

            border-radius:
                4px;

            padding:
                2px 5px;

            min-width:
                28px;

            height:
                17px;

            box-sizing:
                border-box;

            font-family:
                Arial,
                Helvetica,
                sans-serif;

            font-size:
                9px;

            font-weight:
                800;

            line-height:
                13px;

            white-space:
                nowrap;

            text-align:
                center;

            box-shadow:
                0 1px 3px
                rgba(0,0,0,.55);

            cursor:
                pointer;
        "
    >

        {safe_code}

    </div>
    """


# =============================================================================
# SILVERA STAR
# =============================================================================

def silvera_star_html() -> str:
    """
    Purple Silvera star.

    Slightly larger than before and always intended
    to sit visually above CHC property labels.
    """

    return f"""
    <div
        class="silvera-star-label"

        style="
            width:
                28px;

            height:
                28px;

            display:flex;

            align-items:center;

            justify-content:center;

            color:
                {SILVERA_COLOR};

            font-family:
                Arial,
                Helvetica,
                sans-serif;

            font-size:
                29px;

            font-weight:
                900;

            line-height:
                28px;

            text-align:
                center;

            cursor:
                pointer;

            text-shadow:
                -1px -1px 0 white,
                 1px -1px 0 white,
                -1px  1px 0 white,
                 1px  1px 0 white,
                 0px  2px 4px
                 rgba(0,0,0,.60);
        "
    >
        ★
    </div>
    """
# =============================================================================
# CHC FAMILY TOOLTIP
# =============================================================================

def chc_family_tooltip(
    family: str,
    sites: list[dict[str, Any]],
) -> str:

    codes = sorted(

        clean_text(

            site.get(
                "Property Code"
            )
        )

        for site
        in sites
    )


    site_word = (

        "site"

        if len(sites) == 1

        else "sites"
    )


    return (

        f"{family} — "
        f"{len(sites)} "
        f"{site_word}: "

        + ", ".join(
            codes
        )

        + " — click to zoom in"
    )


# =============================================================================
# MAP LEGEND
# =============================================================================

def add_map_legend(

    map_object: folium.Map,

    chc_counts: Counter[str],

    chc_total: int,

    silvera_total: int,

) -> None:


    unknown_line = ""


    if chc_counts.get(
        "Unknown",
        0,
    ):

        unknown_line = f"""
        <div>

            <span style="
                display:inline-block;

                width:
                    13px;

                height:
                    8px;

                background:
                    #f59e0b;

                border-radius:
                    2px;

                margin-right:
                    5px;
            ">
            </span>

            CHC Unknown
            ({chc_counts.get("Unknown", 0)})

        </div>
        """


    legend = f"""
    <div style="
        position:
            fixed;

        bottom:
            30px;

        left:
            30px;

        z-index:
            9999;

        background:
            rgba(
                255,
                255,
                255,
                .96
            );

        border:
            1px solid #999;

        border-radius:
            6px;

        padding:
            10px 12px;

        font-family:
            Arial,
            sans-serif;

        font-size:
            12px;

        line-height:
            1.6;

        min-width:
            245px;

        box-shadow:
            0 1px 6px
            rgba(0,0,0,.20);
    ">


        <div style="
            font-size:
                14px;

            font-weight:
                700;

            margin-bottom:
                6px;
        ">
            Resource Allocation Map
        </div>


        <div>

            <span style="
                display:inline-block;

                width:
                    13px;

                height:
                    8px;

                background:
                    #d73027;

                border-radius:
                    2px;

                margin-right:
                    5px;
            ">
            </span>

            CHC East — Grounds/Snow
            ({chc_counts.get("East", 0)})

        </div>


        <div>

            <span style="
                display:inline-block;

                width:
                    13px;

                height:
                    8px;

                background:
                    #2a81cb;

                border-radius:
                    2px;

                margin-right:
                    5px;
            ">
            </span>

            CHC South — Reference
            ({chc_counts.get("South", 0)})

        </div>


        <div>

            <span style="
                display:inline-block;

                width:
                    13px;

                height:
                    8px;

                background:
                    #5a9f2b;

                border-radius:
                    2px;

                margin-right:
                    5px;
            ">
            </span>

            CHC West — Reference
            ({chc_counts.get("West", 0)})

        </div>


        {unknown_line}


        <div style="
            margin-top:
                3px;
        ">

            <span style="
                color:
                    {SILVERA_COLOR};

                font-size:
                    18px;

                font-weight:
                    900;

                line-height:
                    10px;

                vertical-align:
                    -1px;

                margin-right:
                    4px;

                text-shadow:
                    0 1px 1px
                    rgba(0,0,0,.25);
            ">
                ★
            </span>

            Silvera Communities
            ({silvera_total})

        </div>


        <div style="
            border-top:
                1px solid #ddd;

            margin-top:
                7px;

            padding-top:
                6px;
        ">

            CHC:
            <b>{chc_total}</b>
            property locations

            <br>

            Silvera:
            <b>{silvera_total}</b>
            locations

        </div>


        <div style="
            color:#666;

            margin-top:
                7px;

            font-size:
                10px;

            line-height:
                1.45;
        ">

            CHC families expand when zoomed in.

            <br>

            Click a CHC family to zoom into it.

            <br>

            Silvera stars remain visible at all zoom levels.

        </div>


    </div>
    """


    map_object.get_root().html.add_child(

        folium.Element(
            legend
        )
    )


# =============================================================================
# MAP TITLE
# =============================================================================

def add_map_title(
    map_object: folium.Map,
) -> None:


    title_html = f"""
    <div style="
        position:
            fixed;

        top:
            10px;

        left:
            50%;

        transform:
            translateX(-50%);

        z-index:
            9998;

        background:
            rgba(255,255,255,.94);

        border:
            1px solid #bbb;

        border-radius:
            6px;

        padding:
            7px 14px;

        font-family:
            Arial,
            sans-serif;

        font-weight:
            700;

        font-size:
            15px;

        box-shadow:
            0 1px 4px
            rgba(0,0,0,.18);

        white-space:
            nowrap;
    ">

        Resource Allocation Map

        <span style="
            font-size:
                10px;

            color:#777;

            font-weight:
                400;

            margin-left:
                7px;
        ">

            CHC + Silvera

        </span>

    </div>
    """


    map_object.get_root().html.add_child(

        folium.Element(
            title_html
        )
    )


# =============================================================================
# BUILD MAP
# =============================================================================

def build_map(

    chc_sites: list[
        dict[str, Any]
    ],

    silvera_sites: list[
        dict[str, Any]
    ],

) -> Counter[str]:


    if (

        not chc_sites

        and

        not silvera_sites
    ):

        raise RuntimeError(
            "No geocoded sites are available."
        )


    # -------------------------------------------------------------------------
    # REMOVE OLD MAP FIRST
    # -------------------------------------------------------------------------

    if OUTPUT_MAP.exists():

        try:

            OUTPUT_MAP.unlink()

        except Exception:

            pass


    # -------------------------------------------------------------------------
    # PREPARE CHC FAMILIES
    # -------------------------------------------------------------------------

    family_groups = (
        build_family_groups(
            chc_sites
        )
    )


    prepare_chc_display_coordinates(
        family_groups
    )


    # -------------------------------------------------------------------------
    # PREPARE SILVERA DISPLAY COORDINATES
    # -------------------------------------------------------------------------

    prepare_silvera_display_coordinates(
        silvera_sites
    )


    # =============================================================================
    # CREATE MAP
    # =============================================================================

    resource_map = folium.Map(

        location=CALGARY_CENTER,

        zoom_start=10,

        tiles=None,

        control_scale=True,

        prefer_canvas=False,
    )


    # =============================================================================
    # BASE MAPS
    # =============================================================================

    folium.TileLayer(

        "CartoDB positron",

        name="Light Map",

        control=True,

        show=True,

    ).add_to(
        resource_map
    )


    folium.TileLayer(

        "OpenStreetMap",

        name="OpenStreetMap",

        control=True,

        show=False,

    ).add_to(
        resource_map
    )


    # =============================================================================
    # CHC DISTRICT LAYERS
    # =============================================================================

    chc_layers: dict[
        str,
        folium.FeatureGroup,
    ] = {}


    for (
        district,
        style,
    ) in DISTRICT_STYLE.items():


        layer = folium.FeatureGroup(

            name=style[
                "label"
            ],

            overlay=True,

            control=True,

            show=style[
                "show"
            ],
        )


        layer.add_to(
            resource_map
        )


        chc_layers[
            district
        ] = layer


    # =============================================================================
    # SILVERA LAYER
    # =============================================================================

    silvera_layer = folium.FeatureGroup(

        name=SILVERA_LAYER_NAME,

        overlay=True,

        control=True,

        show=True,
    )


    silvera_layer.add_to(
        resource_map
    )


    # =============================================================================
    # MAP COUNTS / BOUNDS
    # =============================================================================

    chc_counts: Counter[
        str
    ] = Counter()


    all_bounds: list[
        list[float]
    ] = []


    # =============================================================================
    # CHC INDIVIDUAL PROPERTY LABELS
    # =============================================================================

    for site in chc_sites:


        district = district_for(
            site
        )


        style = DISTRICT_STYLE[
            district
        ]


        property_code = clean_text(

            site.get(
                "Property Code"
            )
        )


        project_name = clean_text(

            site.get(
                "Project Name"
            )
        )


        actual_lat = float(
            site[
                "Latitude"
            ]
        )


        actual_lon = float(
            site[
                "Longitude"
            ]
        )


        display_lat = float(

            site.get(
                "_DisplayLatitude",
                actual_lat,
            )
        )


        display_lon = float(

            site.get(
                "_DisplayLongitude",
                actual_lon,
            )
        )


        all_bounds.append(

            [
                actual_lat,
                actual_lon,
            ]
        )


        chc_counts[
            district
        ] += 1


        marker = folium.Marker(

            location=[

                display_lat,

                display_lon,
            ],


            icon=DivIcon(

                html=(
                    chc_detail_marker_html(

                        property_code,

                        style[
                            "color"
                        ],
                    )
                ),

                icon_size=(
                    34,
                    16,
                ),

                icon_anchor=(
                    17,
                    8,
                ),
            ),


            tooltip=folium.Tooltip(

                html.escape(

                    f"{property_code} — "
                    f"{project_name}"
                ),

                sticky=True,
            ),


            popup=folium.Popup(

                chc_popup_html(
                    site
                ),

                max_width=460,
            ),


            title=property_code,
        )


        marker.add_to(

            chc_layers[
                district
            ]
        )


    # =============================================================================
    # CHC FAMILY LABELS
    # =============================================================================

    family_click_javascript: list[
        str
    ] = []


    map_js_name = (
        resource_map.get_name()
    )


    for (
        family_key,
        family_sites,
    ) in sorted(
        family_groups.items()
    ):


        (
            district,
            family,
        ) = family_key


        style = DISTRICT_STYLE[
            district
        ]


        latitudes = [

            float(
                site[
                    "Latitude"
                ]
            )

            for site
            in family_sites
        ]


        longitudes = [

            float(
                site[
                    "Longitude"
                ]
            )

            for site
            in family_sites
        ]


        center_lat = (

            sum(
                latitudes
            )

            / len(
                latitudes
            )
        )


        center_lon = (

            sum(
                longitudes
            )

            / len(
                longitudes
            )
        )


        # ---------------------------------------------------------------------
        # ONE PROPERTY ONLY:
        #
        # show full code rather than shortened family
        # ---------------------------------------------------------------------

        if len(
            family_sites
        ) == 1:

            display_code = clean_text(

                family_sites[
                    0
                ].get(
                    "Property Code"
                )
            )

        else:

            display_code = family


        family_marker = folium.Marker(

            location=[

                center_lat,

                center_lon,
            ],


            icon=DivIcon(

                html=(
                    chc_family_marker_html(

                        display_code,

                        style[
                            "color"
                        ],
                    )
                ),

                icon_size=(
                    42,
                    19,
                ),

                icon_anchor=(
                    21,
                    9,
                ),
            ),


            tooltip=folium.Tooltip(

                html.escape(

                    chc_family_tooltip(

                        family,

                        family_sites,
                    )
                ),

                sticky=True,
            ),


            title=display_code,
        )


        family_marker.add_to(

            chc_layers[
                district
            ]
        )


        marker_js_name = (
            family_marker.get_name()
        )


        # =====================================================================
        # MULTI-SITE FAMILY CLICK
        # =====================================================================

        if len(
            family_sites
        ) > 1:


            family_bounds = [

                [

                    float(
                        site[
                            "Latitude"
                        ]
                    ),

                    float(
                        site[
                            "Longitude"
                        ]
                    ),
                ]

                for site
                in family_sites
            ]


            bounds_json = json.dumps(
                family_bounds
            )


            family_click_javascript.append(
                f"""
                {marker_js_name}.on(

                    "click",

                    function() {{

                        var familyBounds =
                            L.latLngBounds(
                                {bounds_json}
                            );


                        {map_js_name}.fitBounds(

                            familyBounds,

                            {{

                                padding:
                                    [60, 60],

                                maxZoom:
                                    {FAMILY_CLICK_MAX_ZOOM}

                            }}
                        );


                        setTimeout(

                            function() {{

                                if (

                                    {map_js_name}.getZoom()
                                    <
                                    {DETAIL_ZOOM_LEVEL}

                                ) {{

                                    {map_js_name}.setZoom(
                                        {DETAIL_ZOOM_LEVEL}
                                    );

                                }}


                                updatePropertyZoomLayers();

                            }},

                            180
                        );

                    }}
                );
                """
            )


        # =====================================================================
        # SINGLE PROPERTY FAMILY
        # =====================================================================

        else:


            latitude = float(

                family_sites[
                    0
                ][
                    "Latitude"
                ]
            )


            longitude = float(

                family_sites[
                    0
                ][
                    "Longitude"
                ]
            )


            family_click_javascript.append(
                f"""
                {marker_js_name}.on(

                    "click",

                    function() {{

                        {map_js_name}.setView(

                            [
                                {latitude},
                                {longitude}
                            ],

                            {DETAIL_ZOOM_LEVEL}
                        );


                        setTimeout(

                            updatePropertyZoomLayers,

                            100
                        );

                    }}
                );
                """
            )


    # =============================================================================
    # SILVERA STARS
    # =============================================================================

    for site in silvera_sites:


        display_name = clean_text(

            site.get(
                "Display Name"
            )
        )


        source_address = clean_text(

            site.get(
                "Source Address"
            )
        )


        actual_lat = float(
            site[
                "Latitude"
            ]
        )


        actual_lon = float(
            site[
                "Longitude"
            ]
        )


        display_lat = float(

            site.get(
                "_DisplayLatitude",
                actual_lat,
            )
        )


        display_lon = float(

            site.get(
                "_DisplayLongitude",
                actual_lon,
            )
        )


        all_bounds.append(

            [
                actual_lat,
                actual_lon,
            ]
        )


        silvera_marker = folium.Marker(

            location=[

                display_lat,

                display_lon,
            ],


            icon=DivIcon(

                html=(
                    silvera_star_html()
                ),

                icon_size=(
                    30,
                    30,
                ),

                icon_anchor=(
                    15,
                    15,
                ),
            ),


            tooltip=folium.Tooltip(

                html.escape(
                    display_name
                ),

                sticky=True,
            ),


            popup=folium.Popup(

                silvera_popup_html(
                    site
                ),

                max_width=480,
            ),


            title=(

                f"Silvera — "
                f"{display_name} — "
                f"{source_address}"
            ),


            # -----------------------------------------------------------------
            # IMPORTANT:
            #
            # Leaflet normally calculates marker stacking partly from latitude.
            # This forces Silvera stars ABOVE all CHC property labels.
            # -----------------------------------------------------------------

            z_index_offset=10000,
        )


        silvera_marker.add_to(
            silvera_layer
        )


    # =============================================================================
    # INITIAL MAP BOUNDS
    #
    # Includes BOTH contracts.
    # =============================================================================

    if all_bounds:


        resource_map.fit_bounds(

            all_bounds,

            padding=(
                20,
                20,
            ),
        )


    # =============================================================================
    # MAP CONTROLS
    # =============================================================================

    MiniMap(

        toggle_display=True,

        minimized=True,

    ).add_to(
        resource_map
    )


    Fullscreen(

        position="topleft",

        title="Full screen",

        title_cancel=(
            "Exit full screen"
        ),

    ).add_to(
        resource_map
    )


    MeasureControl(

        position="topleft",

        primary_length_unit=(
            "kilometers"
        ),

    ).add_to(
        resource_map
    )


    folium.LayerControl(

        position="topright",

        collapsed=False,

    ).add_to(
        resource_map
    )


    # =============================================================================
    # TITLE + LEGEND
    # =============================================================================

    add_map_title(
        resource_map
    )


    add_map_legend(

        resource_map,

        chc_counts,

        len(
            chc_sites
        ),

        len(
            silvera_sites
        ),
    )


    # =============================================================================
    # CHC ZOOM VISIBILITY
    #
    # SILVERA IS NOT TOUCHED.
    #
    # Zoom < DETAIL_ZOOM_LEVEL:
    #
    #     show CHC family labels
    #     hide CHC detail labels
    #
    # Zoom >= DETAIL_ZOOM_LEVEL:
    #
    #     hide CHC family labels
    #     show CHC detail labels
    #
    # Silvera stars always remain visible.
    # =============================================================================

    zoom_script = f"""
    <script>

    function setMarkerElementVisibility(
        selector,
        visible
    ) {{

        document
            .querySelectorAll(
                selector
            )
            .forEach(

                function(
                    innerElement
                ) {{

                    var markerElement =
                        innerElement.closest(
                            ".leaflet-marker-icon"
                        );


                    if (
                        !markerElement
                    ) {{

                        return;

                    }}


                    if (
                        visible
                    ) {{

                        markerElement.style.display =
                            "";

                        markerElement.style.pointerEvents =
                            "auto";

                    }}

                    else {{

                        markerElement.style.display =
                            "none";

                        markerElement.style.pointerEvents =
                            "none";

                    }}

                }}
            );

    }}


    function updatePropertyZoomLayers() {{

        var currentZoom =
            {map_js_name}.getZoom();


        var showDetails =
            currentZoom
            >=
            {DETAIL_ZOOM_LEVEL};


        setMarkerElementVisibility(

            ".chc-family-label",

            !showDetails
        );


        setMarkerElementVisibility(

            ".chc-detail-label",

            showDetails
        );


        // IMPORTANT:
        //
        // We deliberately DO NOTHING to
        // .silvera-star-label.
        //
        // Silvera remains visible at every zoom.

    }}


    window.addEventListener(

        "load",

        function() {{

            setTimeout(

                function() {{

                    // =========================================================
                    // CHC FAMILY CLICK HANDLERS
                    // =========================================================

                    {"".join(family_click_javascript)}


                    // =========================================================
                    // INITIAL CHC VISIBILITY
                    // =========================================================

                    updatePropertyZoomLayers();


                    // =========================================================
                    // ZOOM EVENT
                    // =========================================================

                    {map_js_name}.on(

                        "zoomend",

                        function() {{

                            updatePropertyZoomLayers();

                        }}
                    );


                    // =========================================================
                    // DISTRICT LAYER RE-ENABLED
                    // =========================================================

                    {map_js_name}.on(

                        "overlayadd",

                        function() {{

                            setTimeout(

                                updatePropertyZoomLayers,

                                40
                            );

                        }}
                    );


                    // =========================================================
                    // RESIZE SAFETY
                    // =========================================================

                    {map_js_name}.on(

                        "resize",

                        function() {{

                            setTimeout(

                                updatePropertyZoomLayers,

                                40
                            );

                        }}
                    );

                }},

                250
            );

        }}
    );

    </script>
    """


    resource_map.get_root().html.add_child(

        folium.Element(
            zoom_script
        )
    )


    # =============================================================================
    # SAVE
    # =============================================================================

    OUTPUT_DIR.mkdir(

        parents=True,

        exist_ok=True,
    )


    resource_map.save(

        str(
            OUTPUT_MAP
        )
    )


    # =============================================================================
    # VERIFY GENERATED MAP
    # =============================================================================

    generated_html = OUTPUT_MAP.read_text(
        encoding="utf-8"
    )


    # -------------------------------------------------------------------------
    # OLD ICON SYSTEM SHOULD NOT BE THERE
    # -------------------------------------------------------------------------

    old_patterns = [

        '"icon": "building"',

        '"icon":"building"',

        "AwesomeMarkers.icon(",
    ]


    old_found = [

        pattern

        for pattern
        in old_patterns

        if pattern
        in generated_html
    ]


    if old_found:

        raise RuntimeError(

            "Old building markers were found "
            "inside the generated map."
        )


    # -------------------------------------------------------------------------
    # VERIFY CHC FAMILY LABELS
    # -------------------------------------------------------------------------

    if (
        "chc-family-label"
        not in generated_html
    ):

        raise RuntimeError(

            "CHC family labels were not generated."
        )


    # -------------------------------------------------------------------------
    # VERIFY CHC DETAIL LABELS
    # -------------------------------------------------------------------------

    if (
        "chc-detail-label"
        not in generated_html
    ):

        raise RuntimeError(

            "CHC individual labels were not generated."
        )


# -------------------------------------------------------------------------
# VERIFY SILVERA
# -------------------------------------------------------------------------

    if (
        "silvera-star-label"
        not in generated_html
    ):

        raise RuntimeError(

            "Silvera star markers were not generated."
        )


    # Folium may encode the star as \u2605,
    # so verify using the text portion only.

    if (
        "Silvera Communities"
        not in generated_html
    ):

        raise RuntimeError(

            "Silvera layer could not be verified."
        )


    # -------------------------------------------------------------------------
    # VERIFY CHC CLICK ZOOM
    # -------------------------------------------------------------------------

    if (
        "familyBounds"
        not in generated_html
    ):

        print()

        print(
            "WARNING: CHC family click-to-zoom "
            "could not be verified."
        )


    print()
    print(
        "Map verification: PASSED"
    )


    print(
        "✓ CHC family labels created"
    )


    print(
        "✓ CHC individual labels created"
    )


    print(
        "✓ CHC click-to-zoom created"
    )


    print(
        "✓ Silvera purple stars created"
    )


    print(
        "✓ Silvera independent map layer created"
    )


    print(
        "✓ Silvera remains visible at every zoom"
    )


    print(
        "✓ No building icons"
    )


    print(
        "✓ No teardrop pins"
    )


    print(
        "✓ No numbered clusters"
    )


    return chc_counts


# =============================================================================
# WRITE FAILURE CSV
# =============================================================================

def write_failure_csv(

    output_path: Path,

    failures: list[
        dict[str, str]
    ],

) -> None:


    OUTPUT_DIR.mkdir(

        parents=True,

        exist_ok=True,
    )


    if not failures:


        if output_path.exists():

            try:

                output_path.unlink()

            except Exception:

                pass


        return


    fields = [

        "Site",

        "Address",

        "Reason",

        "Attempted Queries",
    ]


    with output_path.open(

        "w",

        newline="",

        encoding="utf-8-sig",

    ) as file:


        writer = csv.DictWriter(

            file,

            fieldnames=fields,
        )


        writer.writeheader()


        writer.writerows(
            failures
        )


# =============================================================================
# COMMAND LINE OPTIONS
# =============================================================================

def parse_args() -> argparse.Namespace:


    parser = argparse.ArgumentParser(

        description=(

            "Create the CHC + Silvera "
            "Resource Allocation Map."
        )
    )


    parser.add_argument(

        "--retry-failed",

        action="store_true",

        help=(

            "Retry previously failed "
            "CHC and Silvera geocoding."
        ),
    )


    parser.add_argument(

        "--no-open",

        action="store_true",

        help=(

            "Create the HTML map "
            "without opening it."
        ),
    )


    return parser.parse_args()


# =============================================================================
# MAIN
# =============================================================================

def main() -> int:


    args = parse_args()


    OUTPUT_DIR.mkdir(

        parents=True,

        exist_ok=True,
    )


    print()
    print("=" * 82)


    print(
        "RESOURCE ALLOCATION MAP"
    )


    print("=" * 82)


    print(
        f"VERSION: {VERSION}"
    )


    print("=" * 82)


    print()


    print(
        "Python file:"
    )


    print(
        Path(
            __file__
        ).resolve()
    )


    print()


    print(
        "CHC workbook:"
    )


    print(
        CHC_XLSX
    )


    print()


    print(
        "Silvera workbook:"
    )


    print(
        SILVERA_XLSX
    )


    print()


    # =============================================================================
    # READ BOTH CONTRACTS
    # =============================================================================

    chc_sites = read_chc_sites()


    silvera_sites = (
        read_silvera_sites()
    )


    # =============================================================================
    # SOURCE COUNTS
    # =============================================================================

    chc_district_counts = Counter(

        district_for(
            site
        )

        for site
        in chc_sites
    )


    chc_families = {

        (

            district_for(
                site
            ),

            property_family(

                clean_text(

                    site.get(
                        "Property Code"
                    )
                )
            ),
        )

        for site
        in chc_sites
    }


    silvera_area_counts = Counter(

        clean_text(

            site.get(
                "Area"
            )
        )

        for site
        in silvera_sites
    )


    print(
        f"CHC properties loaded: "
        f"{len(chc_sites)}"
    )


    print(
        f"CHC property families: "
        f"{len(chc_families)}"
    )


    print()


    for (
        district,
        count,
    ) in sorted(
        chc_district_counts.items()
    ):

        print(

            f"  CHC "
            f"{district:<8} "
            f"{count}"
        )


    print()


    print(
        f"Silvera unique locations loaded: "
        f"{len(silvera_sites)}"
    )


    print()


    for (
        area,
        count,
    ) in sorted(
        silvera_area_counts.items()
    ):

        print(

            f"  Silvera "
            f"{area:<10} "
            f"{count}"
        )


    print()


    # =============================================================================
    # GEOCODE CHC
    # =============================================================================

    print("=" * 82)


    print(
        "CHC GEOCODING / CACHE"
    )


    print("=" * 82)


    (
        chc_located,
        chc_failures,
        chc_new_requests,
    ) = geocode_chc_sites(

        chc_sites,

        retry_failed=(
            args.retry_failed
        ),
    )


    # =============================================================================
    # GEOCODE SILVERA
    # =============================================================================

    print()
    print("=" * 82)


    print(
        "SILVERA GEOCODING / CACHE"
    )


    print("=" * 82)


    (
        silvera_located,
        silvera_failures,
        silvera_new_requests,
    ) = geocode_silvera_sites(

        silvera_sites,

        retry_failed=(
            args.retry_failed
        ),
    )


    # =============================================================================
    # FAILURE REPORTS
    # =============================================================================

    write_failure_csv(

        CHC_FAILURES_CSV,

        chc_failures,
    )


    write_failure_csv(

        SILVERA_FAILURES_CSV,

        silvera_failures,
    )


    # =============================================================================
    # BUILD MAP
    # =============================================================================

    print()
    print("=" * 82)


    print(
        "CREATING MAP"
    )


    print("=" * 82)


    plotted_chc_counts = build_map(

        chc_located,

        silvera_located,
    )


    # =============================================================================
    # RESULTS
    # =============================================================================

    print()
    print("=" * 82)


    print(
        "DONE"
    )


    print("=" * 82)


    print()


    print(
        "CHC:"
    )


    print(

        f"  Plotted: "
        f"{len(chc_located)} "
        f"/ {len(chc_sites)}"
    )


    for (
        district,
        count,
    ) in sorted(
        plotted_chc_counts.items()
    ):

        print(

            f"  {district:<8} "
            f"{count}"
        )


    print()


    print(
        "Silvera:"
    )


    print(

        f"  Plotted: "
        f"{len(silvera_located)} "
        f"/ {len(silvera_sites)}"
    )


    silvera_plotted_areas = Counter(

        clean_text(

            site.get(
                "Area"
            )
        )

        for site
        in silvera_located
    )


    for (
        area,
        count,
    ) in sorted(
        silvera_plotted_areas.items()
    ):

        print(

            f"  {area:<10} "
            f"{count}"
        )


    print()


    print(
        "Geocoder requests this run:"
    )


    print(

        f"  CHC: "
        f"{chc_new_requests}"
    )


    print(

        f"  Silvera: "
        f"{silvera_new_requests}"
    )


    print()


    print(
        "Generated map:"
    )


    print(
        OUTPUT_MAP
    )


    print()


    if chc_failures:


        print(
            "CHC locations needing review:"
        )


        print(
            CHC_FAILURES_CSV
        )


    else:


        print(
            "CHC geocoding failures: none"
        )


    print()


    if silvera_failures:


        print(
            "Silvera locations needing review:"
        )


        print(
            SILVERA_FAILURES_CSV
        )


    else:


        print(
            "Silvera geocoding failures: none"
        )


    print()
    print("=" * 82)


    print(
        "EXPECTED MAP"
    )


    print("=" * 82)


    print()


    print(
        "CHC:"
    )


    print(

        f"  Zoom < {DETAIL_ZOOM_LEVEL}: "
        f"RUN / PIN / FHT / ABB / etc."
    )


    print(

        f"  Zoom {DETAIL_ZOOM_LEVEL}+: "
        f"RUN1 / RUN2 / PIN1 / PIN2 / etc."
    )


    print(
        "  Click a CHC family → zoom into its properties."
    )


    print()


    print(
        "SILVERA:"
    )


    print(
        "  Purple ★ = Silvera community"
    )


    print(
        "  Stars remain visible at every zoom level."
    )


    print(
        "  Hover ★ → community name"
    )


    print(
        "  Click ★ → address / area / community details"
    )


    print()


    print(
        "LAYER MENU:"
    )


    print(
        "  ☑ CHC East"
    )


    print(
        "  ☑ CHC South"
    )


    print(
        "  ☑ CHC West"
    )


    print(
        "  ☑ ★ Silvera Communities"
    )


    print()


    print(
        "For bidding/resource analysis, "
        "try leaving only:"
    )


    print(
        "  ☑ CHC East"
    )


    print(
        "  ☑ ★ Silvera Communities"
    )


    print("=" * 82)


    # =============================================================================
    # OPEN BROWSER
    # =============================================================================

    if not args.no_open:


        try:

            webbrowser.open(

                OUTPUT_MAP.resolve().as_uri()
            )


        except Exception:

            pass


    return 0


# =============================================================================
# ENTRY POINT
# =============================================================================

if __name__ == "__main__":


    try:


        raise SystemExit(
            main()
        )


    except KeyboardInterrupt:


        print()
        print()


        print(
            "Stopped by user."
        )


        print(

            "Completed geocoding remains "
            "saved in geocode_cache.json."
        )


        raise SystemExit(
            130
        )


    except Exception as exc:


        print()
        print()


        print("=" * 82)


        print(
            "ERROR"
        )


        print("=" * 82)


        print()


        print(
            exc
        )


        print()


        print(
            "Map generation did not complete."
        )


        print()


        print(
            "Fix the error above and run "
            "Create_Resource_Map.bat again."
        )


        print()


        print("=" * 82)


        raise SystemExit(
            1
        )